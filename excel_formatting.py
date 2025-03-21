from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule, Rule
from openpyxl.styles.differential import DifferentialStyle

# Formatting the final excel file
def format_excel_file(path):
    red_fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")
    bold_font = Font(bold=True)
    dxf = DifferentialStyle(fill=red_fill, font=bold_font)

    wb = load_workbook(path)
    ws = wb.active

    rule =  Rule(type="expression", formula=['ISNUMBER(SEARCH("TimeOut", C1))'], dxf=dxf)

    ws.conditional_formatting.add('C1:C9999', rule)

    ws.auto_filter.ref = f"A1:D1"

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    last_row = ws.max_row
    last_column = ws.max_column
    last_column_letter = ws.cell(row=1, column=last_column).column_letter

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_column):
        for cell in row:
            cell.border = thin_border

    wb.save(path)

if __name__ == "__main__":
    print("Not to be used on it's own, functions must be imported")